#!/usr/bin/env python3
"""
Generate Supplementary Tables for Project 1 Manuscript
------------------------------------------------------
Creates a multi-sheet Excel workbook with all supplementary data.

Spondyloarthritis AI Computational Biology Institute
PI: Pawel Kuklik | Computational Analyst: Perplexity Computer
"""

import pandas as pd
import numpy as np
import json
import os
import warnings
warnings.filterwarnings('ignore')

BASE = '/home/user/workspace/project1_cross_disease_metaanalysis_spa'
OUT = os.path.join(BASE, 'docs', 'Supplementary_Tables.xlsx')

# ============================================================
# Helper to format Excel sheets
# ============================================================
def format_sheet(writer, sheet_name, df, title=None):
    """Write dataframe to sheet with formatting."""
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2 if title else 0)
    ws = writer.sheets[sheet_name]
    if title:
        ws.cell(row=1, column=1, value=title)


# ============================================================
# TABLE S1: Dataset Summary
# ============================================================
print("Generating Table S1: Dataset Summary...")

with open(os.path.join(BASE, 'data', 'metadata', 'final_dataset_inventory.json')) as f:
    inventory = json.load(f)

# DEG counts per dataset
deg_counts = {}
for dataset_info in inventory:
    acc = dataset_info['accession']
    sig_file = os.path.join(BASE, 'results', 'deg', f'{acc}_deg_significant.csv')
    if os.path.exists(sig_file):
        sig = pd.read_csv(sig_file)
        deg_counts[acc] = len(sig) - 1 if len(sig) > 1 else 0  # header row
    else:
        deg_counts[acc] = 'N/A'

s1_data = []
for d in inventory:
    s1_data.append({
        'GEO Accession': d['accession'],
        'Disease': d['disease'],
        'Tissue': d['tissue'],
        'Platform': d['platform'],
        'Data Type': d['data_type'],
        'Total Samples': d['n_samples'],
        'Disease Samples': d['n_disease'],
        'Control Samples': d['n_control'],
        'DEGs (|log2FC|>1, padj<0.05)': deg_counts.get(d['accession'], 'N/A'),
        'Included in Analysis': 'Yes' if d.get('usable', True) else 'No'
    })

s1 = pd.DataFrame(s1_data)

# ============================================================
# TABLE S2: Full DEG Results (top genes per dataset)
# ============================================================
print("Generating Table S2: Differential Expression Results...")

s2_sheets = {}
for dataset_info in inventory:
    acc = dataset_info['accession']
    deg_file = os.path.join(BASE, 'results', 'deg', f'{acc}_deg_results.csv')
    if os.path.exists(deg_file):
        df = pd.read_csv(deg_file)
        if 'gene' in df.columns or 'Gene' in df.columns:
            gene_col = 'gene' if 'gene' in df.columns else 'Gene'
            # Get significant genes or top by p-value
            if 'padj' in df.columns:
                df_sorted = df.sort_values('padj')
            elif 'p_value' in df.columns:
                df_sorted = df.sort_values('p_value')
            else:
                df_sorted = df
            # Take top 500 per dataset to keep file manageable
            s2_sheets[acc] = df_sorted.head(500)

# ============================================================
# TABLE S3: Meta-Analysis Results
# ============================================================
print("Generating Table S3: Meta-Analysis Results...")

meta_as = pd.read_csv(os.path.join(BASE, 'results', 'meta_analysis', 'meta_significant_AS_only.csv'))
meta_cross = os.path.join(BASE, 'results', 'meta_analysis', 'meta_significant_cross_disease.csv')
if os.path.exists(meta_cross):
    meta_cd = pd.read_csv(meta_cross)
else:
    meta_cd = pd.DataFrame()

# Clean up meta-analysis table
meta_as_clean = meta_as[['gene', 'fisher_p', 'fisher_padj', 'n_datasets', 'weighted_log2FC', 
                          'consensus_direction', 'n_up', 'n_down', 'datasets']].copy()
meta_as_clean.columns = ['Gene', 'Fisher p-value', 'Fisher padj', 'N Datasets', 
                          'Weighted log2FC', 'Direction', 'N Up', 'N Down', 'Datasets']
meta_as_clean = meta_as_clean.sort_values('Fisher p-value')

# ============================================================
# TABLE S4: Functional Enrichment
# ============================================================
print("Generating Table S4: Functional Enrichment...")

enrich_file = os.path.join(BASE, 'results', 'enrichment', 'AS_meta_enrichment.csv')
s4 = pd.read_csv(enrich_file)
s4_clean = s4[['source', 'term_id', 'term_name', 'p_value', 'term_size', 
                'intersection_size', 'precision', 'recall']].copy()
s4_clean.columns = ['Source', 'Term ID', 'Term Name', 'p-value', 'Term Size',
                     'Genes in Term', 'Precision', 'Recall']

# ============================================================
# TABLE S5: PPI Network & Hub Genes
# ============================================================
print("Generating Table S5: PPI Network & Hub Genes...")

hub_genes = pd.read_csv(os.path.join(BASE, 'results', 'networks', 'hub_genes.csv'))
hub_clean = hub_genes[['gene', 'degree', 'betweenness', 'closeness', 'composite_rank']].copy()
hub_clean.columns = ['Gene', 'Degree', 'Betweenness Centrality', 'Closeness Centrality', 'Composite Rank']
hub_clean = hub_clean.sort_values('Composite Rank')

ppi_edges = pd.read_csv(os.path.join(BASE, 'results', 'networks', 'ppi_edges.csv'))

# ============================================================
# TABLE S6: WGCNA Modules
# ============================================================
print("Generating Table S6: WGCNA Module Summary...")

# Module-trait correlations
as_trait = pd.read_csv(os.path.join(BASE, 'results', 'wgcna', 'AS_module_trait_cor.csv'))
ibd_trait = pd.read_csv(os.path.join(BASE, 'results', 'wgcna', 'IBD_module_trait_cor.csv'))

# Module gene lists
with open(os.path.join(BASE, 'results', 'wgcna', 'AS_module_gene_lists.json')) as f:
    as_modules = json.load(f)
with open(os.path.join(BASE, 'results', 'wgcna', 'IBD_module_gene_lists.json')) as f:
    ibd_modules = json.load(f)

# Create module summary
wgcna_summary = []
for mod_name, genes in as_modules.items():
    # Find matching trait correlation
    mod_num = mod_name.split('_')[1]
    trait_row = as_trait[as_trait['module'] == int(mod_num)] if mod_num.isdigit() else pd.DataFrame()
    color = mod_name.split('_')[-1] if '_' in mod_name else mod_name
    wgcna_summary.append({
        'Network': 'AS (GSE18781)',
        'Module': f'ME{color}',
        'N Genes': len(genes),
        'Trait Correlation': trait_row['cor_trait'].values[0] if len(trait_row) > 0 else np.nan,
        'Trait p-value': trait_row['p_trait'].values[0] if len(trait_row) > 0 else np.nan,
        'Top 20 Genes': ', '.join(genes[:20])
    })

for mod_name, genes in ibd_modules.items():
    mod_num = mod_name.split('_')[1]
    trait_row = ibd_trait[ibd_trait['module'] == int(mod_num)] if mod_num.isdigit() else pd.DataFrame()
    color = mod_name.split('_')[-1] if '_' in mod_name else mod_name
    wgcna_summary.append({
        'Network': 'IBD (GSE59071)',
        'Module': f'ME{color}',
        'N Genes': len(genes),
        'Trait Correlation': trait_row['cor_trait'].values[0] if len(trait_row) > 0 else np.nan,
        'Trait p-value': trait_row['p_trait'].values[0] if len(trait_row) > 0 else np.nan,
        'Top 20 Genes': ', '.join(genes[:20])
    })

s6 = pd.DataFrame(wgcna_summary)

# Module preservation
pres_as2ibd = pd.read_csv(os.path.join(BASE, 'results', 'wgcna', 'preservation_AS_to_IBD.csv'))
pres_ibd2as = pd.read_csv(os.path.join(BASE, 'results', 'wgcna', 'preservation_IBD_to_AS.csv'))

# Meta-gene overlap
overlap_as = pd.read_csv(os.path.join(BASE, 'results', 'wgcna', 'AS_meta_overlap_genes.csv'))
overlap_ibd = pd.read_csv(os.path.join(BASE, 'results', 'wgcna', 'IBD_meta_overlap_genes.csv'))

# ============================================================
# TABLE S7: Immune Cell Deconvolution
# ============================================================
print("Generating Table S7: Immune Cell Deconvolution...")

deconv_results = []
for disease in ['AS', 'IBD', 'jSpA', 'PsA']:
    sf = os.path.join(BASE, 'results', 'deconvolution', f'{disease}_deconvolution_stats.csv')
    if os.path.exists(sf):
        df = pd.read_csv(sf)
        if len(df) > 0:
            df['Disease'] = disease
            deconv_results.append(df)

if deconv_results:
    s7 = pd.concat(deconv_results, ignore_index=True)
    # Standardize column names
    col_rename = {}
    for c in s7.columns:
        if 'cohens' in c.lower() or 'cohen' in c.lower():
            col_rename[c] = "Cohen's d"
        elif c == 'p_value':
            col_rename[c] = 'p-value'
        elif c == 'p_adjusted':
            col_rename[c] = 'FDR'
    s7 = s7.rename(columns=col_rename)
else:
    s7 = pd.DataFrame()

# Cross-disease Kruskal-Wallis
kw_file = os.path.join(BASE, 'results', 'deconvolution', 'cross_disease_kruskal_updated.csv')
if os.path.exists(kw_file):
    kw = pd.read_csv(kw_file)
else:
    kw = pd.DataFrame()

# ============================================================
# TABLE S8: Drug-Gene Interactions
# ============================================================
print("Generating Table S8: Drug-Gene Interactions...")

drug = pd.read_csv(os.path.join(BASE, 'results', 'drug_repurposing', 'drug_gene_interactions.csv'))
drug_clean = drug.copy()
drug_clean.columns = [c.replace('_', ' ').title() for c in drug_clean.columns]

# ============================================================
# WRITE ALL TABLES TO EXCEL
# ============================================================
print("\nWriting to Excel workbook...")

with pd.ExcelWriter(OUT, engine='openpyxl') as writer:
    # S1: Dataset Summary
    s1.to_excel(writer, sheet_name='S1_Datasets', index=False, startrow=1)
    ws = writer.sheets['S1_Datasets']
    ws.cell(row=1, column=1, value='Supplementary Table S1. Summary of GEO datasets included in the cross-disease transcriptomic meta-analysis')
    
    # S2: DEG results per dataset (separate sub-sheets or combined)
    if s2_sheets:
        # Combined DEG table with dataset column
        all_degs = []
        for acc, df in s2_sheets.items():
            df_copy = df.head(100).copy()  # Top 100 per dataset
            df_copy.insert(0, 'Dataset', acc)
            # Find disease from inventory
            disease = next((d['disease'] for d in inventory if d['accession'] == acc), 'Unknown')
            df_copy.insert(1, 'Disease', disease)
            all_degs.append(df_copy)
        s2_combined = pd.concat(all_degs, ignore_index=True)
        s2_combined.to_excel(writer, sheet_name='S2_DEGs', index=False, startrow=1)
        ws = writer.sheets['S2_DEGs']
        ws.cell(row=1, column=1, value='Supplementary Table S2. Top differentially expressed genes per dataset (ranked by adjusted p-value)')
    
    # S3: Meta-analysis
    meta_as_clean.to_excel(writer, sheet_name='S3_MetaAnalysis', index=False, startrow=1)
    ws = writer.sheets['S3_MetaAnalysis']
    ws.cell(row=1, column=1, value='Supplementary Table S3. AS meta-analysis results (40 genes significant across ≥2 datasets)')
    
    # S4: Enrichment
    s4_clean.to_excel(writer, sheet_name='S4_Enrichment', index=False, startrow=1)
    ws = writer.sheets['S4_Enrichment']
    ws.cell(row=1, column=1, value='Supplementary Table S4. Functional enrichment analysis of AS meta-significant genes (g:Profiler)')
    
    # S5: PPI & Hub Genes
    hub_clean.to_excel(writer, sheet_name='S5_HubGenes', index=False, startrow=1)
    ws = writer.sheets['S5_HubGenes']
    ws.cell(row=1, column=1, value='Supplementary Table S5a. PPI network hub gene centrality metrics')
    
    ppi_edges.to_excel(writer, sheet_name='S5b_PPIEdges', index=False, startrow=1)
    ws = writer.sheets['S5b_PPIEdges']
    ws.cell(row=1, column=1, value='Supplementary Table S5b. PPI network edges (STRING, score ≥ 0.4)')
    
    # S6: WGCNA
    s6.to_excel(writer, sheet_name='S6_WGCNA', index=False, startrow=1)
    ws = writer.sheets['S6_WGCNA']
    ws.cell(row=1, column=1, value='Supplementary Table S6a. WGCNA module summary with trait correlations')
    
    pres_as2ibd.to_excel(writer, sheet_name='S6b_Preservation', index=False, startrow=1)
    ws = writer.sheets['S6b_Preservation']
    ws.cell(row=1, column=1, value='Supplementary Table S6b. Module preservation statistics (AS→IBD)')
    
    if not pres_ibd2as.empty:
        # Add IBD→AS below
        startrow = len(pres_as2ibd) + 4
        pres_ibd2as.to_excel(writer, sheet_name='S6b_Preservation', index=False, startrow=startrow)
        ws.cell(row=startrow, column=1, value='Module preservation statistics (IBD→AS)')
    
    if not overlap_as.empty:
        overlap_combined = pd.concat([
            overlap_as.assign(Network='AS'),
            overlap_ibd.assign(Network='IBD') if not overlap_ibd.empty else pd.DataFrame()
        ], ignore_index=True)
        overlap_combined.to_excel(writer, sheet_name='S6c_MetaGeneOverlap', index=False, startrow=1)
        ws = writer.sheets['S6c_MetaGeneOverlap']
        ws.cell(row=1, column=1, value='Supplementary Table S6c. AS meta-significant genes found in WGCNA modules')
    
    # S7: Deconvolution
    if not s7.empty:
        s7.to_excel(writer, sheet_name='S7_Deconvolution', index=False, startrow=1)
        ws = writer.sheets['S7_Deconvolution']
        ws.cell(row=1, column=1, value='Supplementary Table S7a. Immune cell deconvolution results per disease (ssGSEA)')
    
    if not kw.empty:
        kw.to_excel(writer, sheet_name='S7b_CrossDisease', index=False, startrow=1)
        ws = writer.sheets['S7b_CrossDisease']
        ws.cell(row=1, column=1, value='Supplementary Table S7b. Cross-disease Kruskal-Wallis test for immune cell type variation')
    
    # S8: Drug-gene interactions
    drug_clean.to_excel(writer, sheet_name='S8_DrugGene', index=False, startrow=1)
    ws = writer.sheets['S8_DrugGene']
    ws.cell(row=1, column=1, value='Supplementary Table S8. Drug-gene interactions for AS meta-significant genes (DGIdb)')

print(f"\nSupplementary tables saved: {OUT}")

# Print sheet summary
import openpyxl
wb = openpyxl.load_workbook(OUT)
print(f"\nWorkbook contains {len(wb.sheetnames)} sheets:")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  {name}: {ws.max_row} rows × {ws.max_column} columns")
